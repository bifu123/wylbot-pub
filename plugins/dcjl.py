'''
登记群中的政务大厅督查反馈
'''
import aiohttp
import asyncio
from models_load import *



################# 主函数 ##################
def fun_dcjl(name_space, function_type, post_type, user_state:list, priority, role=[], block=False):
    def decorator(func):
        func._name_space = name_space
        func._function_type = function_type
        func._post_type = post_type
        func._priority = priority
        func._user_state = user_state
        func._role = role
        func._block = block
        return func
    return decorator




################# 子函数 ##################
# 插件函数示例1
@fun_dcjl(name_space="督查记录", function_type="parallel", post_type="message", user_state=["插件问答"], priority=3, role=[])
def write_to_excel(data):
    import os
    import datetime
    import openpyxl
    try:
        import openpyxl
    except:
        os.system('pip install openpyxl')
        import openpyxl

    try:
        source_id = f'{data["group_id"]}'
    except:
        source_id = f'{data["user_id"]}'
    user_id = f'{data["user_id"]}'
    message = data["message"]

    # 打开或创建xlsx文件
    if os.path.exists("督查记录.xlsx"):
        wb = openpyxl.load_workbook("督查记录.xlsx")
    else:
        wb = openpyxl.Workbook()
        wb.active.append(["id", "记录人", "记录来源", "督查记事", "记录时间"])

    # 获取当前活动的工作表
    ws = wb.active

    # 获取当前日期时间
    write_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 获取当前行数，即 id 值
    row_id = ws.max_row + 1

    # 将数据写入工作表
    # 将数据写入工作表
    ws.append([row_id, user_id, source_id, message, write_time])

    # 保存xlsx文件
    wb.save("督查记录.xlsx")

    return f"请这样回答：'已记录，新的ID为：{row_id}'，不要再添加其它的推理和词句"





