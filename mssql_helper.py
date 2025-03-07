import pymssql
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import datetime

class DatabaseHandler:
    def __init__(self, server, database, user, password, charset='utf8'):
        self.server = server
        self.database = database
        self.user = user
        self.password = password
        self.charset = charset
        self.conn = None

    def connect(self):
        self.conn = pymssql.connect(server=self.server, database=self.database, user=self.user, password=self.password, charset=self.charset)

    def disconnect(self):
        if self.conn:
            self.conn.close()

    def execute_command(self, sql, params=None):
        self.connect()
        cursor = self.conn.cursor()
        cursor.execute(sql, params)
        self.conn.commit()
        self.disconnect()

    def execute_select_all(self, sql):
        self.connect()
        cursor = self.conn.cursor()
        cursor.execute(sql)
        results = cursor.fetchall()
        self.disconnect()
        return results if results else None

    def execute_select_one(self, sql):
        self.connect()
        cursor = self.conn.cursor()
        cursor.execute(sql)
        result = cursor.fetchone()
        self.disconnect()
        return result if result else None

# 初始化数据库处理器
db_handler = DatabaseHandler(server='192.168.66.6', database='car', user='sa', password='Shift962512')

# 执行不返回结果的sql
def command_handler(sql, params=None):
    db_handler.execute_command(sql, params)

# 执行返回结果的查询（多条记录）
def select_handler_all(sql):
    return db_handler.execute_select_all(sql)

# 执行返回结果的查询（一条记录）
def select_handler_one(sql):
    return db_handler.execute_select_one(sql)


# 插入当前聊天历史记录
def insert_chat_history_to_mssql(source_id, user_name, content, user_state, name_space=""):
    sql = "INSERT INTO history_now (source_id, [user], [content], user_state, name_space) VALUES (%s, %s, %s, %s, %s)"
    params = (source_id, user_name, content, user_state, name_space)
    db_handler.execute_command(sql, params)


# # 写入聊天的记录到EXCEL
# def insert_chat_history_to_excel(source_id, user, content,user_state="聊天", name_space="test"):
#     # 检查文件是否存在
#     filename = 'history.xlsx'
#     if not os.path.isfile(filename):
#         # 如果文件不存在，创建新文件并写入表头
#         wb = Workbook()
#         ws = wb.active
#         ws.append(["source_id", "user", "content", "create_time", "user_state", "name_space"])
#         wb.save(filename)

#     # 打开工作簿并插入新记录
#     wb = load_workbook(filename)
#     ws = wb.active
#     ws.append([source_id, user, content, datetime.datetime.now(), user_state, name_space])
#     wb.save(filename)
    
# # 从当前聊天历史记录表中删除时间最晚的1条记录
# def delete_oldest_records(source_id, user_state, name_space=""):
#     sql = '''DELETE FROM history_now 
#                  WHERE source_id = %s 
#                  AND user_state = %s 
#                  AND name_space = %s
#                  AND id in (SELECT TOP 1 id FROM history_now 
# 				     WHERE source_id = %s 
#                      AND user_state = %s 
#                      AND name_space = %s ORDER BY timestamp)'''
#     params = (source_id, user_state, name_space, source_id, user_state, name_space)
#     db_handler.execute_command(sql, params)

# # 从当前聊天历史记录表中删除符合条件的所有记录
# def delete_all_records(source_id, user_state, name_space):
#     sql = "DELETE FROM history_now WHERE source_id = %s and user_state = %s and name_space = %s"
#     params = (source_id, user_state, name_space)
#     db_handler.execute_command(sql, params)

# # 从数据库中提取聊天历史记录
# def fetch_chat_history(source_id, user_state, name_space):
#     sql = "SELECT [user], [content] content FROM history_now WHERE source_id = %s and user_state = %s and name_space = %s ORDER BY timestamp"
#     params = (source_id, user_state, name_space)
#     try:
#         return select_handler_all(sql, params)
#     except:
#         return []

